#!/usr/bin/env python3
"""
Convert an ASKCOS Tree Builder export (list of networkx node-link trees) into an
Excel workbook (or CSVs) summarising routes, steps, and chemicals.

Routes sheet carries three parallel one-cell views of each pathway:
  - "formulae pathway"   : each step as  reactants -> product  using molecular formulae
  - "names pathway"      : same, using compound names (falls back to formula when no
                           name is available; names require --online or an internal source)
  - "structures pathway" : a vertical reaction scheme (backbone drawn; co-reactants as
                           text labels on the arrows). Full structures per compound live
                           on the Chemicals sheet.

Everything chemical (formula, MW, InChIKey) is derived locally with RDKit.
NAME and CAS are NOT in the file and cannot come from structure; they are blank unless
--online is passed (PubChem lookup by InChIKey).

  !! --online sends your structures to PubChem. Do NOT use it for proprietary targets. !!

Usage
-----
    python treejson_to_xlsx.py treeResults.json                 # -> treeResults.xlsx
    python treejson_to_xlsx.py treeResults.json -o routes.xlsx
    python treejson_to_xlsx.py treeResults.json --csv-dir out/  # CSVs (no embedded images)
    python treejson_to_xlsx.py treeResults.json --reactions     # add unique-Reactions sheet
    python treejson_to_xlsx.py treeResults.json --chem-images    # thumbnails on Chemicals sheet
    python treejson_to_xlsx.py treeResults.json --no-structures  # drop the structures column
    python treejson_to_xlsx.py treeResults.json --online         # names+CAS via PubChem (NON-proprietary)
"""

import argparse
import collections
import csv
import io
import json
import os
import re
import sys

from rdkit import Chem, RDLogger
from rdkit.Chem import Descriptors, rdMolDescriptors

RDLogger.logger().setLevel(RDLogger.CRITICAL)


# ---------------------------------------------------------------------------
# Local (RDKit) helpers
# ---------------------------------------------------------------------------

def _mol(smiles):
    return Chem.MolFromSmiles(smiles) if smiles else None


def _canon(smiles):
    m = _mol(smiles)
    return Chem.MolToSmiles(m) if m else smiles


def _formula(smiles):
    m = _mol(smiles)
    return rdMolDescriptors.CalcMolFormula(m) if m else "(parse fail)"


def _heavy(smiles):
    m = _mol(smiles)
    return m.GetNumHeavyAtoms() if m else 0


def props(smiles):
    """Return (canonical_smiles, formula, mol_weight, inchikey, num_carbons)."""
    m = _mol(smiles)
    if m is None:
        return smiles, "(parse fail)", None, "", None
    try:
        canon = Chem.MolToSmiles(m)
    except Exception:
        canon = smiles
    try:
        formula = rdMolDescriptors.CalcMolFormula(m)
    except Exception:
        formula = ""
    try:
        mw = round(Descriptors.MolWt(m), 2)
    except Exception:
        mw = None
    try:
        ikey = Chem.MolToInchiKey(m)
    except Exception:
        ikey = ""
    ncarb = sum(1 for a in m.GetAtoms() if a.GetAtomicNum() == 6)
    return canon, formula, mw, ikey, ncarb


def _stoich(token, n):
    """Coefficient in front, chemistry-style: '3 CH4O' (or just 'CH4O' when n==1)."""
    return f"{n} {token}" if n > 1 else token


def reactant_formulae(reactant_smiles_list):
    """'3 CH4O, C2H4Cl4Si' — dedup with counts, preserving first-seen order."""
    counts = collections.Counter(reactant_smiles_list)
    parts = []
    for smi in dict.fromkeys(reactant_smiles_list):
        parts.append(_stoich(_formula(smi), counts[smi]))
    return ", ".join(parts)


# ---------------------------------------------------------------------------
# Tree parsing
# ---------------------------------------------------------------------------

def parse_tree(tree):
    """Return root smiles, ordered reaction steps, per-chem roles, and the
    backbone (main-scaffold chain + per-arrow co-reactant text labels)."""
    nodes = {n["id"]: n for n in tree["nodes"]}
    out_deg = collections.Counter(e["from"] for e in tree["edges"])
    in_deg = collections.Counter(e["to"] for e in tree["edges"])
    children = collections.defaultdict(list)
    for e in tree["edges"]:
        children[e["from"]].append(e["to"])

    chem_ids = [i for i, n in nodes.items() if n["type"] == "chemical"]
    rxn_ids = [i for i, n in nodes.items() if n["type"] == "reaction"]

    roots = [i for i in chem_ids if in_deg[i] == 0]
    root = roots[0] if roots else (chem_ids[0] if chem_ids else None)
    root_smiles = nodes[root]["smiles"] if root else ""

    # reaction "step" = number of reaction nodes on the path from root to it
    step_of = {}
    if root is not None:
        stack = [(root, 0)]
        while stack:
            nid, rc = stack.pop()
            if nodes[nid]["type"] == "reaction":
                rc += 1
                step_of[nid] = rc
            for c in children.get(nid, []):
                stack.append((c, rc))

    steps = []
    for rid in rxn_ids:
        smi = nodes[rid]["smiles"]
        left, _, right = smi.partition(">>")
        steps.append({
            "step": step_of.get(rid, None),
            "reaction_smiles": smi,
            "product_smiles": right,
            "reactants": [s for s in left.split(".") if s],
        })
    steps.sort(key=lambda s: (s["step"] if s["step"] is not None else 99, s["product_smiles"]))

    chem_role = {}
    for cid in chem_ids:
        smi = nodes[cid]["smiles"]
        if cid == root:
            chem_role[smi] = "target"
        elif out_deg[cid] == 0:
            chem_role[smi] = "starting_material"
        else:
            chem_role[smi] = "intermediate"

    # backbone: follow the largest-scaffold reactant at each step; co-reactants
    # (from the reaction SMILES, so stoichiometry is correct) become arrow labels.
    chain, labels, cur = [nodes[root]["smiles"]] if root else [""], [""], root
    while cur is not None:
        rx = [c for c in children.get(cur, []) if nodes[c]["type"] == "reaction"]
        if not rx:
            break
        rxn_smiles = nodes[rx[0]]["smiles"]
        reacts = [c for c in children.get(rx[0], []) if nodes[c]["type"] == "chemical"]
        if not reacts:
            break
        cont = max(reacts, key=lambda c: _heavy(nodes[c]["smiles"]))
        chain.append(nodes[cont]["smiles"])
        labels.append(_coreactant_label(rxn_smiles, nodes[cont]["smiles"]))
        cur = cont
    backbone_fwd = list(reversed(chain))
    backbone_lab = list(reversed(labels))   # arrow AFTER fwd[k] carries lab[k]

    return {
        "root_smiles": root_smiles, "graph": tree.get("graph", {}),
        "steps": steps, "chem_role": chem_role,
        "num_chemicals": len(chem_ids),
        "num_starting_materials": sum(1 for r in chem_role.values() if r == "starting_material"),
        "backbone_fwd": backbone_fwd, "backbone_lab": backbone_lab,
        # main-scaffold feedstock: the ultimate starting material on the backbone
        # (backbone_fwd is forward/synthesis order, so index 0 is the deepest SM).
        "feedstock_smiles": backbone_fwd[0] if backbone_fwd else "",
    }


def _coreactant_label(rxn_smiles, continuation_smiles):
    """Reaction-SMILES reactants minus one instance of the backbone continuation,
    as formulae with stoichiometric counts. e.g. 'CH4O x3'."""
    reactants = [r for r in rxn_smiles.split(">>")[0].split(".") if r]
    cc = _canon(continuation_smiles)
    removed, rest = False, []
    for r in reactants:
        if not removed and _canon(r) == cc:
            removed = True
            continue
        rest.append(r)
    counts = collections.Counter(_formula(r) for r in rest)
    order = list(dict.fromkeys(_formula(r) for r in rest))
    return " + ".join(_stoich(f, counts[f]) for f in order)


# ---------------------------------------------------------------------------
# One-cell pathway summaries (forward / synthesis order)
# ---------------------------------------------------------------------------

def _forward_step(product_smiles, reactants, tok):
    counts = collections.Counter(tok(s) for s in reactants)
    order = list(dict.fromkeys(tok(s) for s in reactants))
    left = " + ".join(_stoich(t, counts[t]) for t in order)
    return f"{left} \u2192 {tok(product_smiles)}"


def summary(steps, tok):
    ordered = sorted(steps, key=lambda s: (-(s["step"] if s["step"] is not None else -1),
                                           s["product_smiles"]))
    return "\n".join(f"{i}. {_forward_step(s['product_smiles'], s['reactants'], tok)}"
                     for i, s in enumerate(ordered, start=1))


# ---------------------------------------------------------------------------
# Vertical route scheme image (backbone structures, co-reactants as text)
# ---------------------------------------------------------------------------

def route_image_png(fwd, labels, mw=150, mh=105, gap=34, labelw=160):
    from PIL import Image, ImageDraw, ImageFont
    from rdkit.Chem.Draw import rdMolDraw2D

    def cell(smi):
        m = _mol(smi)
        d = rdMolDraw2D.MolDraw2DCairo(mw, mh)
        if m is not None:
            d.DrawMolecule(m)
        d.FinishDrawing()
        return Image.open(io.BytesIO(d.GetDrawingText())).convert("RGB")

    font = ImageFont.load_default()
    imgs = [cell(s) for s in fwd]
    W, H = mw + labelw, len(imgs) * mh + gap * (len(imgs) - 1)
    canvas = Image.new("RGB", (W, H), "white")
    d = ImageDraw.Draw(canvas)
    y = 0
    for k, im in enumerate(imgs):
        canvas.paste(im, (0, y))
        if k < len(imgs) - 1:
            ax, y0, y1 = mw // 2, y + mh + 2, y + mh + gap - 2
            d.line([(ax, y0), (ax, y1 - 8)], fill="black", width=2)
            d.polygon([(ax, y1), (ax - 5, y1 - 9), (ax + 5, y1 - 9)], fill="black")
            if labels[k]:
                d.text((ax + 10, (y0 + y1) // 2), labels[k], fill="gray", font=font, anchor="lm")
        y += mh + gap
    buf = io.BytesIO()
    canvas.save(buf, format="PNG")
    return buf.getvalue(), W, H


# ---------------------------------------------------------------------------
# Build tables
# ---------------------------------------------------------------------------

ROUTE_GRAPH_COLS = [
    "depth", "num_reactions", "precursor_cost", "atom_economy",
    "score", "avg_score", "min_score", "first_step_score",
    "avg_plausibility", "min_plausibility", "first_step_plausibility", "cluster_id",
]

COL_FORMULAE = "formulae pathway"
COL_NAMES = "names pathway"
COL_STRUCTURES = "structures pathway"


def build_tables(trees, online=False):
    routes, steps_rows = [], []
    chem = {}
    reactions = {}

    def chem_key(smiles):
        canon, formula, mw, ikey, ncarb = props(smiles)
        key = ikey or canon
        if key not in chem:
            chem[key] = {
                "canon": canon, "smiles": canon, "formula": formula, "mol_weight": mw,
                "inchikey": ikey, "num_carbons": ncarb, "name": "", "cas": "",
                "as_target": 0, "as_intermediate": 0, "as_starting_material": 0,
                "as_product": 0, "as_reactant": 0, "pathways": set(),
            }
        return key

    for idx, tree in enumerate(trees, start=1):
        p = parse_tree(tree)
        g = p["graph"]
        _, tform, tmw, _, _ = props(p["root_smiles"])
        _, fform, fmw, _, _ = props(p["feedstock_smiles"])

        route = {
            "pathway_id": idx, "target": p["root_smiles"],
            "target_formula": tform, "target_mol_weight": tmw,
            "feedstock": p["feedstock_smiles"],
            "feedstock_formula": fform, "feedstock_mol_weight": fmw,
            "num_chemicals": p["num_chemicals"],
            "num_starting_materials": p["num_starting_materials"],
            COL_FORMULAE: summary(p["steps"], _formula),
            COL_NAMES: "",          # filled after (optional) name lookup
            COL_STRUCTURES: "",     # image embedded in writer
            "_steps": p["steps"],
            "_backbone": (p["backbone_fwd"], p["backbone_lab"]),
        }
        for c in ROUTE_GRAPH_COLS:
            route[c] = g.get(c, None)
        routes.append(route)

        for s in p["steps"]:
            _, pform, _, _, _ = props(s["product_smiles"])
            steps_rows.append({
                "pathway_id": idx, "step": s["step"],
                "reaction_smiles": s["reaction_smiles"],
                "product_smiles": s["product_smiles"], "product_formula": pform,
                "reactants_smiles": " + ".join(s["reactants"]),
                "reactants_formulae": reactant_formulae(s["reactants"]),
                "num_reactants": len(s["reactants"]),
            })
            rec = reactions.setdefault(s["reaction_smiles"], {
                "reaction_smiles": s["reaction_smiles"], "product_smiles": s["product_smiles"],
                "product_formula": pform, "reactants_formulae": reactant_formulae(s["reactants"]),
                "num_reactants": len(s["reactants"]), "pathways": set(),
            })
            rec["pathways"].add(idx)
            chem[chem_key(s["product_smiles"])]["as_product"] += 1
            for rc in s["reactants"]:
                chem[chem_key(rc)]["as_reactant"] += 1

        for smi, role in p["chem_role"].items():
            k = chem_key(smi)
            chem[k]["pathways"].add(idx)
            chem[k]["as_" + role] += 1

    # chemicals rows
    chem_rows = []
    for rec in chem.values():
        if rec["as_target"] > 0:
            primary = "target"
        elif rec["as_intermediate"] > 0 and rec["as_starting_material"] > 0:
            primary = "intermediate/starting_material"
        elif rec["as_intermediate"] > 0:
            primary = "intermediate"
        elif rec["as_starting_material"] > 0:
            primary = "starting_material"
        else:
            primary = ""
        chem_rows.append({
            "smiles": rec["smiles"], "name": rec["name"], "formula": rec["formula"],
            "mol_weight": rec["mol_weight"], "inchikey": rec["inchikey"], "cas": rec["cas"],
            "num_carbons": rec["num_carbons"], "primary_role": primary,
            "times_as_target": rec["as_target"], "times_as_intermediate": rec["as_intermediate"],
            "times_as_starting_material": rec["as_starting_material"],
            "times_as_product": rec["as_product"], "times_as_reactant": rec["as_reactant"],
            "num_pathways": len(rec["pathways"]),
        })
    chem_rows.sort(key=lambda r: (-r["num_pathways"], str(r["formula"])))

    reaction_rows = []
    for rec in reactions.values():
        reaction_rows.append({k: rec[k] for k in
                              ("reaction_smiles", "product_smiles", "product_formula",
                               "reactants_formulae", "num_reactants")}
                             | {"num_pathways": len(rec["pathways"])})
    reaction_rows.sort(key=lambda r: (-r["num_pathways"], r["reaction_smiles"]))

    if online:
        _fill_names_and_cas(chem_rows)

    # names pathway (uses names where available, else formula)
    name_map = {}
    for r in chem_rows:
        if r["name"]:
            name_map[r["smiles"]] = r["name"]

    def name_tok(smiles):
        return name_map.get(_canon(smiles)) or _formula(smiles)

    for route in routes:
        route[COL_NAMES] = summary(route["_steps"], name_tok)

    return routes, steps_rows, chem_rows, reaction_rows


# ---------------------------------------------------------------------------
# Optional online enrichment (name + CAS) — PubChem by InChIKey. Not proprietary-safe.
# ---------------------------------------------------------------------------

CACHE_FILE = ".pubchem_cache.json"
CAS_RE = re.compile(r"^\d{2,7}-\d{2}-\d$")


def _fill_names_and_cas(chem_rows):
    import time
    import urllib.request

    sys.stderr.write("\n*** --online: sending InChIKeys to PubChem. "
                     "Do NOT use for proprietary compounds. ***\n\n")
    cache = {}
    if os.path.exists(CACHE_FILE):
        try:
            cache = json.load(open(CACHE_FILE))
        except Exception:
            cache = {}
    base = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/inchikey/"

    def fetch(ik):
        if ik in cache:
            return cache[ik]
        name, cas = "", ""
        try:
            with urllib.request.urlopen(base + ik + "/synonyms/JSON", timeout=20) as r:
                syns = json.load(r)["InformationList"]["Information"][0].get("Synonym", [])
            if syns:
                name = syns[0]
            for s in syns:
                if CAS_RE.match(s.strip()):
                    cas = s.strip()
                    break
        except Exception:
            pass
        cache[ik] = {"name": name, "cas": cas}
        time.sleep(0.25)
        return cache[ik]

    for row in chem_rows:
        if row["inchikey"]:
            info = fetch(row["inchikey"])
            row["name"], row["cas"] = info.get("name", ""), info.get("cas", "")
    try:
        json.dump(cache, open(CACHE_FILE, "w"))
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def routes_columns(no_structures):
    cols = ["pathway_id", "target", "target_formula", "target_mol_weight",
            "feedstock", "feedstock_formula", "feedstock_mol_weight",
            COL_FORMULAE, COL_NAMES]
    if not no_structures:
        cols.append(COL_STRUCTURES)
    cols += ["depth", "num_reactions", "precursor_cost", "atom_economy",
             "num_chemicals", "num_starting_materials",
             "score", "avg_score", "min_score", "first_step_score",
             "avg_plausibility", "min_plausibility", "first_step_plausibility", "cluster_id"]
    return cols


STEP_COLS = ["pathway_id", "step", "reaction_smiles", "product_smiles",
             "product_formula", "reactants_smiles", "reactants_formulae", "num_reactants"]
CHEM_COLS = ["smiles", "name", "formula", "mol_weight", "inchikey", "cas", "num_carbons",
             "primary_role", "times_as_target", "times_as_intermediate",
             "times_as_starting_material", "times_as_product", "times_as_reactant", "num_pathways"]
RXN_COLS = ["reaction_smiles", "product_smiles", "product_formula",
            "reactants_formulae", "num_reactants", "num_pathways"]


def write_csvs(csv_dir, routes, steps, chems, reactions, include_reactions, no_structures):
    os.makedirs(csv_dir, exist_ok=True)
    rcols = [c for c in routes_columns(no_structures) if c != COL_STRUCTURES]  # no images in csv
    tables = [("routes", rcols, routes), ("steps", STEP_COLS, steps),
              ("chemicals", CHEM_COLS, chems)]
    if include_reactions:
        tables.append(("reactions", RXN_COLS, reactions))
    for name, cols, rows in tables:
        path = os.path.join(csv_dir, name + ".csv")
        with open(path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            for row in rows:
                w.writerow(row)
        print(f"  -> {path}")


def write_xlsx(path, routes, steps, chems, reactions, include_reactions,
               no_structures, chem_images):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4A6FA5")
    body_font = Font(name="Arial")
    float_cols = {"atom_economy", "score", "avg_score", "min_score", "first_step_score",
                  "avg_plausibility", "min_plausibility", "first_step_plausibility",
                  "target_mol_weight", "feedstock_mol_weight", "mol_weight"}
    text_cols = {COL_FORMULAE, COL_NAMES}

    sheets = [("Routes", routes_columns(no_structures), routes),
              ("Steps", STEP_COLS, steps), ("Chemicals", CHEM_COLS, chems)]
    if include_reactions:
        sheets.append(("Reactions", RXN_COLS, reactions))

    for name, cols, rows in sheets:
        ws = wb.create_sheet(name)
        for j, col in enumerate(cols, start=1):
            c = ws.cell(row=1, column=j, value=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(vertical="center")
        for i, row in enumerate(rows, start=2):
            row_pt = None
            for j, col in enumerate(cols, start=1):
                v = row.get(col, None)
                c = ws.cell(row=i, column=j, value=v)
                c.font = body_font
                if col in float_cols and isinstance(v, float):
                    c.number_format = "0.00" if col in ("mol_weight", "target_mol_weight", "feedstock_mol_weight") else "0.000"
                if col in text_cols and isinstance(v, str) and v:
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                    row_pt = max(row_pt or 0, 14 * (v.count("\n") + 1))
            if name == "Routes" and not no_structures and row.get("_backbone"):
                img_pt = _embed_route_image(ws, i, cols, row)
                if img_pt:
                    row_pt = max(row_pt or 0, img_pt)
            if row_pt:
                ws.row_dimensions[i].height = min(row_pt, 900)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
        for j, col in enumerate(cols, start=1):
            width = max(len(str(col)) + 2, 12)
            if "smiles" in str(col):
                width = 34
            if col in text_cols:
                width = 46
            if col == COL_STRUCTURES:
                width = 44
            if col in ("reactants_formulae", "primary_role", "name"):
                width = 24
            ws.column_dimensions[get_column_letter(j)].width = width

    if chem_images:
        _embed_chem_thumbs(wb["Chemicals"], chems)

    wb.save(path)
    print(f"  -> {path}")


def _embed_route_image(ws, row_idx, cols, route):
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    try:
        fwd, lab = route["_backbone"]
        if not fwd:
            return None
        png, w, h = route_image_png(fwd, lab)
        col_letter = get_column_letter(cols.index(COL_STRUCTURES) + 1)
        img = XLImage(io.BytesIO(png))
        ws.add_image(img, f"{col_letter}{row_idx}")
        return h * 0.75  # px -> points
    except Exception as e:
        sys.stderr.write(f"  (route image failed for pathway {route.get('pathway_id')}: {e})\n")
        return None


def _embed_chem_thumbs(ws, chems):
    from openpyxl.drawing.image import Image as XLImage
    from rdkit.Chem.Draw import rdMolDraw2D
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="structure")
    ws.column_dimensions["A"].width = 22
    for i, row in enumerate(chems, start=2):
        m = _mol(row["smiles"])
        if m is None:
            continue
        try:
            d = rdMolDraw2D.MolDraw2DCairo(150, 110)
            d.DrawMolecule(m)
            d.FinishDrawing()
            ws.add_image(XLImage(io.BytesIO(d.GetDrawingText())), f"A{i}")
            ws.row_dimensions[i].height = 85
        except Exception:
            pass


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("json", help="Tree Builder export (list of node-link trees)")
    ap.add_argument("-o", "--output", help="Output .xlsx path (default: <input>.xlsx)")
    ap.add_argument("--csv-dir", help="Also write CSVs to this directory (no embedded images)")
    ap.add_argument("--no-xlsx", action="store_true", help="Skip the xlsx (use with --csv-dir)")
    ap.add_argument("--reactions", action="store_true", help="Include a unique-Reactions sheet/csv")
    ap.add_argument("--no-structures", action="store_true",
                    help="Drop the 'structures pathway' column (no route images)")
    ap.add_argument("--chem-images", action="store_true",
                    help="Embed per-compound thumbnails on the Chemicals sheet")
    ap.add_argument("--online", action="store_true",
                    help="Fill names+CAS via PubChem. NOT for proprietary structures.")
    args = ap.parse_args()

    with open(args.json) as f:
        trees = json.load(f)
    if not isinstance(trees, list):
        sys.exit("Expected a JSON list of trees (networkx node-link format).")

    print(f"Loaded {len(trees)} trees from {args.json}")
    routes, steps, chems, reactions = build_tables(trees, online=args.online)
    print(f"  {len(routes)} routes | {len(steps)} steps | {len(chems)} unique chemicals "
          f"| {len(reactions)} unique reactions")

    if args.csv_dir:
        write_csvs(args.csv_dir, routes, steps, chems, reactions, args.reactions, args.no_structures)
    if not args.no_xlsx:
        out = args.output or (os.path.splitext(args.json)[0] + ".xlsx")
        write_xlsx(out, routes, steps, chems, reactions, args.reactions,
                   args.no_structures, args.chem_images)


if __name__ == "__main__":
    main()
